#!/usr/bin/env python3
"""Parse the client's real spreadsheets into prisma/real-data.json.

Sources (relative to the repo parent directory):
  - data/decrypted/Engagement contract Tracker _ CRM.xlsx  -> mandates
  - data/decrypted/Tasks Tracker Whatsapp 2026_ CRM.xlsx   -> tasks

Output: prisma/real-data.json  { "mandates": [...], "tasks": [...] }
"""

import datetime
import json
import re
import sys
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")  # openpyxl extension warnings

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent


def _resolve_data_dir():
    """The decrypted spreadsheets live at <repo-parent>/decrypted (current layout)
    or <repo-parent>/data/decrypted (older layout). Use whichever exists."""
    for cand in (REPO_ROOT.parent / "decrypted", REPO_ROOT.parent / "data" / "decrypted"):
        if cand.is_dir():
            return cand
    return REPO_ROOT.parent / "decrypted"


DATA_DIR = _resolve_data_dir()
ENGAGEMENT_XLSX = DATA_DIR / "Engagement contract Tracker _ CRM.xlsx"
TASKS_XLSX = DATA_DIR / "Tasks Tracker Whatsapp 2026_ CRM.xlsx"
INVESTOR_XLSX = DATA_DIR / "Investor Tracker _ CRM.xlsx"
OUT_PATH = REPO_ROOT / "prisma" / "real-data.json"

RECENT_CUTOFF = datetime.datetime(2025, 1, 1)


def norm_name(s):
    """casefold + collapse whitespace, for dedup keys."""
    return re.sub(r"\s+", " ", str(s)).strip().casefold()


def clean_str(v):
    if v is None:
        return None
    s = re.sub(r"\s+", " ", str(v)).strip()
    return s or None


def parse_date(v):
    """Return ISO date string or None. Handles datetimes plus the sheet's
    stringly-typed strays like '17/042025', ' 17/09/25', '14/5/2026'."""
    if isinstance(v, datetime.datetime):
        # guard against spreadsheet junk years
        if 2000 <= v.year <= 2030:
            return v.date().isoformat()
        return None
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
            try:
                d = datetime.datetime.strptime(s, fmt)
                if 2000 <= d.year <= 2030:
                    return d.date().isoformat()
            except ValueError:
                pass
        # digit-run heuristic: '17/042025' -> 17042025 (ddmmyyyy), '170925' (ddmmyy)
        digits = re.sub(r"\D", "", s)
        try:
            if len(digits) == 8:
                d = datetime.datetime.strptime(digits, "%d%m%Y")
            elif len(digits) == 6:
                d = datetime.datetime.strptime(digits, "%d%m%y")
            else:
                return None
            if 2000 <= d.year <= 2030:
                return d.date().isoformat()
        except ValueError:
            return None
    return None


def looks_like_junk_client(name):
    """Section headers / numeric noise in the Client column."""
    if re.fullmatch(r"[\d\s.,/–-]+", name):  # purely numeric-ish
        return True
    low = name.casefold()
    if low in {"client", "clients", "total", "n/a"}:
        return True
    if "tracker" in low and len(name) < 40:
        return True
    return False


def parse_mandates():
    wb = openpyxl.load_workbook(ENGAGEMENT_XLSX, data_only=True)
    ws = wb["Engagement Contract Tracker"]

    raw_rows = 0
    skipped_no_client = 0
    skipped_junk = 0
    deduped = {}  # norm name -> (sort_key, mandate dict)

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=10, values_only=True):
        sno, date, client, nda_sent, nda_signed, ea_sent, ea_signed, lead, source, deal = row
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        client = clean_str(client)
        if not client:
            skipped_no_client += 1
            continue
        if looks_like_junk_client(client):
            skipped_junk += 1
            continue
        raw_rows += 1

        d = parse_date(date)
        nda_s, nda_g = parse_date(nda_sent), parse_date(nda_signed)
        ea_s, ea_g = parse_date(ea_sent), parse_date(ea_signed)

        nda_status = "Signed" if nda_g else ("Sent" if nda_s else "NotSent")
        ea_status = "Signed" if ea_g else ("Sent" if ea_s else "NotSent")
        if ea_g:
            stage = "Signed"
        elif ea_s:
            stage = "Negotiation"
        elif nda_g:
            stage = "Proposal"
        elif nda_s:
            stage = "Qualification"
        else:
            stage = "NewLead"

        all_dates = [x for x in (d, nda_s, nda_g, ea_s, ea_g) if x]
        recent = bool(all_dates) and max(all_dates) >= RECENT_CUTOFF.date().isoformat()

        mandate = {
            "clientName": client,
            "date": d,
            "ndaSentDate": nda_s,
            "ndaSignedDate": nda_g,
            "eaSentDate": ea_s,
            "eaSignedDate": ea_g,
            "ndaStatus": nda_status,
            "eaStatus": ea_status,
            "stage": stage,
            "leadName": clean_str(lead),
            "sourceReferee": clean_str(source),
            "dealInfo": clean_str(deal),
            "recent": recent,
        }

        # Dedup: keep the row with the most recent Date; tie -> most filled fields.
        filled = sum(1 for v in mandate.values() if v)
        sort_key = (d or "0000-00-00", filled)
        key = norm_name(client)
        if key not in deduped or sort_key > deduped[key][0]:
            deduped[key] = (sort_key, mandate)

    mandates = [m for _, m in deduped.values()]
    return mandates, raw_rows, skipped_no_client, skipped_junk


WEEK_DIVIDER_RE = re.compile(
    r"^\s*\d{1,2}\s*(st|nd|rd|th)?\s+?\w*\s*[-–]\s*\d{1,2}.*\d{4}\s*$", re.IGNORECASE
)

STATUS_MAP = {"done": "Done", "ongoing": "Ongoing", "pending": "Pending"}


def parse_tasks():
    wb = openpyxl.load_workbook(TASKS_XLSX, data_only=True)
    ws = wb["Task Tracker"]

    tasks = []
    dividers = 0
    empty = 0
    skipped_no_action = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=8, values_only=True):
        project, action, source, status, deadline, owner, assist, notes = row
        if all(v is None or str(v).strip() == "" for v in row):
            empty += 1
            continue
        project_s = clean_str(project)
        action_s = clean_str(action)
        if project_s and not action_s and WEEK_DIVIDER_RE.match(project_s):
            dividers += 1
            continue
        if not action_s:
            skipped_no_action += 1
            continue

        status_s = STATUS_MAP.get(str(status).strip().casefold() if status else "", "NotStarted")
        tasks.append(
            {
                "project": project_s,
                "actionPoint": action_s,
                "status": status_s,
                "deadline": parse_date(deadline),
                "ownerName": clean_str(owner),
                "assistName": clean_str(assist),
                "notes": clean_str(notes),
            }
        )
    return tasks, dividers, empty, skipped_no_action


# ─────────────────────────────────────────────────────────────────────────────
# Investors, investor contacts, service providers (law firms), partners/referees
# ─────────────────────────────────────────────────────────────────────────────

# Free-text sector prose -> Sector[] enum. Keyword => enum value(s). Best-effort;
# the full raw text is always preserved in the investor's investmentMandate field.
SECTOR_KEYWORDS = [
    (r"agri|farm|crop|agro", ["Agribusiness"]),
    (r"microfinanc|\bmfi\b|sacco|financial|fintech|insur|lending|\bloan|credit|banking|\bbank\b", ["FinancialServices"]),
    (r"fintech|insutech|insurtech", ["Technology", "FinancialServices"]),
    (r"\bict\b|software|telecom|digital|\btech\b|technolog|\bit services", ["Technology"]),
    (r"renewable|solar|wind|hydro|geothermal|biomass", ["RenewableEnergy"]),
    (r"energy|power", ["Energy"]),
    (r"health|pharma|medical|hospital|healthcare", ["Healthcare"]),
    (r"manufactur", ["Manufacturing"]),
    (r"educat|edtech|school", ["Education"]),
    (r"fmcg|consumer goods", ["FMCG"]),
    (r"transport|logistic", ["TransportLogistics"]),
    (r"real estate|proptech|property", ["RealEstate"]),
    (r"infrastructure", ["Infrastructure"]),
    (r"water|sanitation", ["WaterSanitation"]),
    (r"hospitality|tourism", ["Hospitality"]),
]

GEO_KEYWORDS = [
    (r"east africa", "EastAfrica"),
    (r"west africa", "WestAfrica"),
    (r"south(ern)? africa", "SouthernAfrica"),
    (r"north africa", "NorthAfrica"),
    (r"sub[- ]?saharan|\bssa\b", "SubSaharanAfrica"),
    (r"pan[- ]?africa", "PanAfrica"),
    (r"\bmena\b", "MENA"),
    (r"europe", "Europe"),
    (r"\busa\b|united states|north america", "USA"),
    (r"global|worldwide", "Global"),
    (r"\bafrica\b", "PanAfrica"),  # bare "Africa" last, so specifics win
]

# Plain substrings; infer_investor_type wraps each in \b…\b word boundaries, so
# short tokens ("bii", "bio", "deg") won't match inside longer words ("biomass").
DFI_NAMES = (
    "ifc", "norfund", "fmo", "cdc", "proparco", "dfc", "bii", "finnfund", "deg",
    "swedfund", "bio", "oikocredit", "incofin", "responsability", "blueorchard",
    "afdb", "afrexim", "symbiotics", "grassroots business fund", "triodos",
)


def map_sectors(text):
    if not text:
        return []
    low = text.casefold()
    out = []
    for pat, vals in SECTOR_KEYWORDS:
        if re.search(pat, low):
            for v in vals:
                if v not in out:
                    out.append(v)
    return out


def map_geographies(text):
    if not text:
        return []
    low = text.casefold()
    out = []
    for pat, val in GEO_KEYWORDS:
        if re.search(pat, low) and val not in out:
            out.append(val)
    # if a specific African region matched, drop the bare PanAfrica fallback
    if len(out) > 1 and "PanAfrica" in out and any(
        v in out for v in ("EastAfrica", "WestAfrica", "SouthernAfrica", "NorthAfrica", "SubSaharanAfrica")
    ):
        out = [v for v in out if v != "PanAfrica"]
    return out


def infer_investor_type(name):
    low = (name or "").casefold()
    if any(re.search(rf"\b{re.escape(d)}\b", low) for d in DFI_NAMES):
        return "DFI"
    if "ventures" in low or re.search(r"\bvc\b", low):
        return "VentureCapital"
    if "family office" in low:
        return "FamilyOffice"
    if "angel" in low:
        return "Angel"
    if re.search(r"\bbank\b", low):
        return "DebtProvider"
    return "PrivateEquity"  # sheet is "Contacts VC PE DFI"; PE is the safe default


def clean_org_name(name):
    """Firm names carry embedded addresses/newlines/notes — keep the first line."""
    if not name:
        return None
    first = str(name).splitlines()[0]
    first = re.sub(r"\s+", " ", first).strip()
    return first or None


def split_contact(raw):
    """'Fidaa Haddad - MD' -> ('Fidaa', 'Haddad', 'MD'). Role after first -/–."""
    if not raw:
        return None
    s = re.sub(r"\s+", " ", str(raw)).strip()
    if not s or s.casefold() == "none":
        return None
    role = None
    m = re.split(r"\s[-–]\s?", s, maxsplit=1)
    if len(m) == 2:
        s, role = m[0].strip(), m[1].strip() or None
    parts = s.split(" ", 1)
    first = parts[0]
    last = parts[1] if len(parts) > 1 else None
    if not first:
        return None
    return {"firstName": first, "lastName": last, "jobTitle": role}


def parse_investors():
    """Investor Tracker 'Contacts VC PE DFI' -> investors with grouped contacts.
    Firm name only appears on the first row of each firm's block; forward-fill."""
    wb = openpyxl.load_workbook(INVESTOR_XLSX, data_only=True, read_only=True)
    ws = wb["Contacts VC PE DFI"]

    investors = []
    cur = None
    for row in ws.iter_rows(min_row=3, max_col=9, values_only=True):
        _sno, firm, website, person, email, tel, geo, sector, others = row
        firm_c = clean_org_name(firm)
        if firm_c and not looks_like_junk_client(firm_c):
            cur = {
                "name": firm_c,
                "investorType": infer_investor_type(firm_c),
                "website": clean_str(website),
                "sectorFocus": map_sectors(clean_str(sector)),
                "geographicFocus": map_geographies(clean_str(geo)),
                "investmentMandate": clean_str(sector),  # preserve raw prose
                "notes": clean_str(others),
                "contacts": [],
            }
            investors.append(cur)
        elif firm_c:
            # a junk firm-name row (section header etc.) ends the current block,
            # so following contacts don't misattach to the previous firm.
            cur = None
        if cur is None:
            continue
        c = split_contact(person)
        if c:
            e = clean_str(email)
            c["email"] = None if (e and e.casefold() == "none") else e
            c["phone"] = clean_str(tel)
            cur["contacts"].append(c)

    # mark first contact of each investor as primary
    for inv in investors:
        if inv["contacts"]:
            inv["contacts"][0]["isPrimary"] = True
    return investors


def extract_email_phone(blob):
    """Pull the first email and first phone-looking run out of a messy cell."""
    if not blob:
        return None, None
    m = re.search(r"[\w.+-]+@[\w.-]+\.\w+", blob)
    email = m.group(0) if m else None
    rest = blob.replace(email, " ") if email else blob
    pm = re.search(r"[+\d][\d\s()+\-]{6,}\d", rest)
    phone = re.sub(r"\s+", " ", pm.group(0)).strip() if pm else None
    return email, phone


def parse_service_providers():
    """Investor Tracker 'Law Firms' sheet -> ServiceProvider(type=LawFirm)."""
    wb = openpyxl.load_workbook(INVESTOR_XLSX, data_only=True, read_only=True)
    if "Law Firms" not in wb.sheetnames:
        return []
    ws = wb["Law Firms"]
    providers = []
    cur = None
    for row in ws.iter_rows(min_row=3, max_col=7, values_only=True):
        _sno, firm, person, contacts, amount, profile, status = row
        firm_c = clean_org_name(firm)
        cval = clean_str(contacts)
        email, phone = extract_email_phone(cval)
        if firm_c and not looks_like_junk_client(firm_c):
            cur = {
                "name": firm_c,
                "type": "LawFirm",
                "contactPerson": clean_str(person),
                "email": email,
                "phone": phone,
                "profile": clean_str(profile),
                "status": clean_str(status),
            }
            providers.append(cur)
        elif firm_c:
            # junk firm-name row ends the current block
            cur = None
        elif cur is not None:
            # fill gaps from continuation rows
            if email and not cur["email"]:
                cur["email"] = email
            if phone and not cur["phone"]:
                cur["phone"] = phone
            if not cur["contactPerson"]:
                cur["contactPerson"] = clean_str(person)
    return providers


# Referral/Source values that are internal staff or status noise, not partners.
INTERNAL_FIRST_NAMES = {
    "amos", "ken", "duncan", "brenda", "brian", "cliff", "james", "evans",
    "sheilla", "susan", "amos g", "evans m", "evans w",
}
STATUS_NOISE = {
    "on hold", "completed", "dropped", "drop", "closed", "to small", "too small",
    "startup", "n/a", "na", "complete", "pending", "done", "ongoing",
}


def clean_referee(raw):
    """Return a cleaned partner name, or None if the value is staff/status noise."""
    s = re.sub(r"\s+", " ", str(raw)).strip()
    if not s:
        return None
    base = re.sub(r"\s*\([^)]*\)\s*$", "", s).strip()  # drop trailing "(Completed)" etc.
    if not base:
        return None
    low = base.casefold()
    if low in STATUS_NOISE or low in INTERNAL_FIRST_NAMES:
        return None
    # phrases like "Dropped too small", "To small", "On hold - ..."
    if any(low.startswith(p) for p in ("dropped", "drop ", "on hold", "closed", "completed")):
        return None
    if "too small" in low or "to small" in low:
        return None
    parts = [p.strip() for p in re.split(r"[/,]", low) if p.strip()]
    if parts and all(p in INTERNAL_FIRST_NAMES for p in parts):
        return None
    return base


def advisor_type_for(name):
    low = name.casefold()
    if re.search(r"\blaw\b|advocat|\bllp\b|khanna|bowmans|legal", low):
        return "Lawyer"
    if re.search(r"advisor|advisory", low):
        return "AdvisoryFirm"
    if re.search(r"consult|accountant|\bcpa\b", low):
        return "Consultant"
    return "Other"


def partner_type_for(name):
    low = name.casefold()
    if re.search(r"\blaw\b|advocat|\bllp\b|legal|khanna|bowmans", low):
        return "LawFirm"
    if re.search(r"consult|advisor|advisory|accountant|capital|investment", low):
        return "Consulting"
    return None


def parse_partners():
    """Engagement Contract Tracker 'Source/Referee' column -> referral partners,
    accumulating the client names each referred (for referredMandates linking)."""
    wb = openpyxl.load_workbook(ENGAGEMENT_XLSX, data_only=True)
    ws = wb["Engagement Contract Tracker"]
    by_name = {}  # norm -> dict
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=10, values_only=True):
        client = clean_str(row[2])
        src = row[8] if len(row) > 8 else None
        if not client or not src:
            continue
        name = clean_referee(src)
        if not name:
            continue
        key = norm_name(name)
        if key not in by_name:
            by_name[key] = {
                "name": name,
                "advisorType": advisor_type_for(name),
                "partnerType": partner_type_for(name),
                "internalOnly": True,
                "referredClients": [],
            }
        rc = by_name[key]["referredClients"]
        if client not in rc:
            rc.append(client)
    return list(by_name.values())


def main():
    mandates, raw_rows, no_client, junk = parse_mandates()
    tasks, dividers, empty, no_action = parse_tasks()
    investors = parse_investors()
    service_providers = parse_service_providers()
    partners = parse_partners()

    OUT_PATH.write_text(json.dumps({
        "mandates": mandates,
        "tasks": tasks,
        "investors": investors,
        "serviceProviders": service_providers,
        "partners": partners,
    }, indent=2))

    recent = sum(1 for m in mandates if m["recent"])
    stages = {}
    for m in mandates:
        stages[m["stage"]] = stages.get(m["stage"], 0) + 1
    statuses = {}
    for t in tasks:
        statuses[t["status"]] = statuses.get(t["status"], 0) + 1

    inv_contacts = sum(len(i["contacts"]) for i in investors)
    partner_links = sum(len(p["referredClients"]) for p in partners)

    print(f"Engagement tracker: {raw_rows} client rows "
          f"(skipped: {no_client} no-client, {junk} junk)")
    print(f"  -> {len(mandates)} deduped mandates ({recent} recent >= {RECENT_CUTOFF.date()})")
    print(f"  stages: {stages}")
    print(f"Task tracker: {len(tasks)} tasks "
          f"(skipped: {dividers} week dividers, {empty} empty, {no_action} without action point)")
    print(f"  statuses: {statuses}")
    print(f"Investor tracker: {len(investors)} firms, {inv_contacts} contacts")
    print(f"Service providers (law firms): {len(service_providers)}")
    print(f"Partners/referees: {len(partners)} ({partner_links} referral links)")
    print(f"Wrote {OUT_PATH}")


if __name__ == "__main__":
    sys.exit(main())
